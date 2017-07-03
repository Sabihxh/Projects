from datetime import datetime
from openpyxl import load_workbook
import string
from flpy.excel import rows_to_excel

uppercase_alphabets = string.ascii_uppercase


wb = load_workbook(filename = './excel/testtest.xlsx', read_only=True)
ws = wb['Downloaded']

result = []

for cell_index_letter in range(0, 26):
    cell_letter = uppercase_alphabets[cell_index_letter]
    cell_index = '%s1' % cell_letter
    c = ws[cell_index]
    if c:
        for y in range(2, 31):
            column = '%s%s' % (cell_index[0], y)
            if ws[column].value:
                print ws[column].value


